from odoo import models,fields,api
import json
import openpyxl


import logging
_logger = logging.getLogger(__name__)

class ProductTemplate(models.Model):
    _inherit = "product.template"

    raisin_type_id = fields.Many2one(
        'raisin.type',
        string="Raisin Type",
        required=True
    )
    is_raisin_category = fields.Boolean(
        string="Is Raisin Category",
        compute="_compute_is_raisin_category",
        store=True
    )
    @api.depends('categ_id.is_raisin')
    def _compute_is_raisin_category(self):
        for rec in self:
            rec.is_raisin_category = rec.categ_id.is_raisin
            
    @api.onchange('categ_id')
    def _onchange_categ_id(self):
        if self.categ_id and self.categ_id.is_raisin:
            if not self.raisin_type_id:
                self.raisin_type_id = self.categ_id.raisin_type_id
        elif not self._origin or self.raisin_type_id:
            self.raisin_type_id = False

class ProductCategory(models.Model):
    _inherit = "product.category"
    
    is_raisin = fields.Boolean(string="Is Raisin")
    raisin_type_id = fields.Many2one("raisin.type", string="Default Raisin Type")
    
    # Direct file upload field
    template_file = fields.Binary(
        string="Upload Template (Excel/CSV)",
        help="Upload your pre-made calculation sheet"
    )
    template_filename = fields.Char(string="Template Filename")
    
    # Spreadsheet template as a separate record
    template_spreadsheet_id = fields.Many2one(
        'crm.lead.spreadsheet',
        string="Template Spreadsheet",
        ondelete='set null'
    )
    
    has_template = fields.Boolean(
        string="Has Template", 
        compute='_compute_has_template'
    )
    
    def _compute_has_template(self):
        for category in self:
            # SAFER APPROACH: Use try-catch for template_spreadsheet_id access
            try:
                has_spreadsheet = bool(category.template_spreadsheet_id)
            except Exception:
                has_spreadsheet = False
                
            has_file = bool(category.template_file)
            category.has_template = has_spreadsheet or has_file
    
    def action_upload_and_create_template(self):
        """Upload Excel and convert to spreadsheet template"""
        self.ensure_one()
        
        if not self.template_file:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'warning',
                    'message': 'Please upload a file first!',
                }
            }
        
        try:
            # Create new spreadsheet template
            new_spreadsheet = self.env['crm.lead.spreadsheet'].create({
                'name': f'{self.name} - Calculation Template',
                'category_id': self.id,
            })
            
            # Convert uploaded file to spreadsheet data
            spreadsheet_data = self._convert_file_to_spreadsheet(self.template_file, self.template_filename)
            if spreadsheet_data:
                new_spreadsheet.raw_spreadsheet_data = json.dumps(spreadsheet_data)
            
            # Link it to category
            self.template_spreadsheet_id = new_spreadsheet
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'success',
                    'message': f'Template uploaded successfully for {self.name}!',
                }
            }
        except Exception as e:
            _logger.error("Error creating template: %s", str(e))
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'error',
                    'message': f'Error creating template: {str(e)}',
                }
            }
    
    def _convert_file_to_spreadsheet(self, file_data, filename):
        """Convert Excel/CSV to Odoo spreadsheet format"""
        import base64
        from io import BytesIO
        
        if not file_data or not filename:
            return None
            
        try:
            # Decode file
            file_content = base64.b64decode(file_data)
            
            # Basic spreadsheet structure
            spreadsheet_data = {
                'version': 1,
                'sheets': [{
                    'id': 'template_sheet',
                    'name': 'Template',
                    'colNumber': 26,
                    'rowNumber': 100,
                    'cells': {},
                    'merges': []
                }]
            }
            
            # Parse Excel file
            if filename.lower().endswith(('.xlsx', '.xls')):
                try:
                    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
                    sheet = workbook.active
                    
                    cells = {}
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                col_letter = openpyxl.utils.get_column_letter(cell.column)
                                cell_ref = f"{col_letter}{cell.row}"
                                
                                # Store cell data
                                cell_content = str(cell.value)
                                cells[cell_ref] = {
                                    'content': cell_content,
                                }
                    
                    spreadsheet_data['sheets'][0]['cells'] = cells
                    
                except ImportError:
                    _logger.warning("openpyxl not installed, skipping Excel parsing")
                    return None
                    
            elif filename.lower().endswith('.csv'):
                # CSV parsing
                try:
                    import csv
                    csv_content = file_content.decode('utf-8')
                    reader = csv.reader(csv_content.splitlines())
                    
                    cells = {}
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, value in enumerate(row, start=1):
                            if value and value.strip():
                                col_letter = self._get_column_letter(col_idx)
                                cell_ref = f"{col_letter}{row_idx}"
                                cells[cell_ref] = {'content': value.strip()}
                    
                    spreadsheet_data['sheets'][0]['cells'] = cells
                    
                except Exception as e:
                    _logger.error("Error parsing CSV: %s", str(e))
                    return None
            
            return spreadsheet_data
            
        except Exception as e:
            _logger.error("Error converting file to spreadsheet: %s", str(e))
            return None
    
    def _get_column_letter(self, col_idx):
        """Convert column index to letter (A, B, C, ... Z, AA, AB, etc.)"""
        letters = ''
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters
    
    def action_open_template(self):
        """Open existing template spreadsheet"""
        self.ensure_one()
        if self.template_spreadsheet_id and self.template_spreadsheet_id.exists():
            return self.template_spreadsheet_id.action_open_spreadsheet()
        else:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'warning',
                    'message': 'No template spreadsheet found for this category',
                }
            }
    
    def action_create_blank_template(self):
        """Create a blank template spreadsheet"""
        self.ensure_one()
        
        new_spreadsheet = self.env['crm.lead.spreadsheet'].create({
            'name': f'{self.name} - Blank Template',
            'category_id': self.id,
        })
        
        self.template_spreadsheet_id = new_spreadsheet
        
        return new_spreadsheet.action_open_spreadsheet()